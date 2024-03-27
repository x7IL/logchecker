import re
import csv
import sys
import asyncio
import aiohttp
import socket
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


class AuthLogParser:
    def __init__(self, log_file):
        # Initializes parser with log file and sets up caches and configurations.
        self.log_file = log_file
        self.domain_name_cache = {}
        self.geolocation_cache = {}
        # IP ranges for local networks.
        self.local_networks = [
            ("10.0.0.0", "10.255.255.255"),
            ("172.16.0.0", "172.31.255.255"),
            ("192.168.0.0", "192.168.255.255"),
        ]
        # Headers for the output Excel report.
        self.headers = [
            "IP Address",
            "Domain Name",
            "Country",
            "Region",
            "City",
            "Localisation",
            "Organization",
            "Postal",
            "Timezone",
            "Start Time",
            "End Time",
            "Successful Attempts",
            "Failed Attempts",
            "Total Attempts",
            "Malicious/Not Sure/No",
            "Impacted Users",
            "Invalid Users",
            "Ports",
            "Success Details",
        ]

        self.malicious_threshold = 5  # Minimum number of attempts to flag an activity as potentially malicious.
        self.batch_size = 100  # Number of IP addresses to process in parallel during domain and geolocation lookups.
        self.timeout = 5  # Timeout for HTTP requests in seconds.

    # Checks if an IP is local.
    def is_local_ip(self, ip_address):
        # Check if the IP address is within local network ranges.
        ip_int = self.ip_to_int(ip_address)
        return any(
            ip_int >= self.ip_to_int(start) and ip_int <= self.ip_to_int(end)
            for start, end in self.local_networks
        )

    # Converts an IP string to an integer.
    def ip_to_int(self, ip):
        # Convert IP address string to an integer.
        parts = ip.split(".")
        return (
            (int(parts[0]) << 24)
            | (int(parts[1]) << 16)
            | (int(parts[2]) << 8)
            | int(parts[3])
        )

    # Fetches domain name asynchronously.
    async def get_domain_name(self, ip_address):
        # Asynchronously retrieves the domain name for an IP address, caching the result.
        if ip_address in self.domain_name_cache or self.is_local_ip(ip_address):
            return
        try:
            # Execute gethostbyaddr to prevent blocking.
            hostname, _, _ = await asyncio.wait_for(
                asyncio.get_event_loop().run_in_executor(
                    None, socket.gethostbyaddr, ip_address
                ),
                timeout=self.timeout,
            )
            self.domain_name_cache[ip_address] = hostname
        except Exception:
            self.domain_name_cache[ip_address] = "N/A"

    # Fetches geolocation data asynchronously.
    async def geolocate_ip(self, ip_address, session):
        if ip_address in self.geolocation_cache or self.is_local_ip(ip_address):
            return
        try:
            async with session.get(
                f"https://ipinfo.io/{ip_address}/json", timeout=self.timeout
            ) as response:
                data = await response.json()
                self.geolocation_cache[ip_address] = {
                    "country": data.get("country", "N/A"),
                    "region": data.get("region", "N/A"),
                    "city": data.get("city", "N/A"),
                    "loc": data.get("loc", "N/A"),
                    "org": data.get("org", "N/A"),
                    "postal": data.get("postal", "N/A"),
                    "timezone": data.get("timezone", "N/A"),
                }
        except Exception as e:
            self.geolocation_cache[ip_address] = {
                "country": "N/A",
                "region": "N/A",
                "city": "N/A",
                "loc": "N/A",
                "org": "N/A",
                "postal": "N/A",
                "timezone": "N/A",
                "error": str(e),
            }

    # Processes IP addresses in batches for domain and geolocation data.
    async def resolve_addresses_batched(self, ip_addresses):
        # Resolves domain names and geolocations in batches for efficiency.
        async with aiohttp.ClientSession() as session:
            tasks = []
            for ip_address in ip_addresses:
                if ip_address not in self.domain_name_cache and not self.is_local_ip(
                    ip_address
                ):
                    tasks.append(self.get_domain_name(ip_address))
                if ip_address not in self.geolocation_cache and not self.is_local_ip(
                    ip_address
                ):
                    tasks.append(self.geolocate_ip(ip_address, session))
            await asyncio.gather(*tasks)

    # Parses the log file.
    def parse_auth_log(self):
        # Initialize storage structures correctly.
        attacks, sudo_usage, other_activities, logs_by_command = {}, {}, [], {}
        current_year = datetime.now().year

        with open(self.log_file, "r") as file:
            for line in file:
                date_match = re.search(r"^\w{3} \d{1,2} \d{2}:\d{2}:\d{2}", line)
                if not date_match:
                    continue

                date_str = f"{date_match.group(0)} {current_year}"
                date_time = datetime.strptime(date_str, "%b %d %H:%M:%S %Y")
                date_time_str = date_time.strftime("%Y-%m-%d %H:%M:%S")

                ip_match = re.search(r"from (\d+\.\d+\.\d+\.\d+)", line)
                if ip_match:
                    self.process_sshd_line(ip_match, line, date_time, attacks)

                sudo_match = re.search(
                    r"sudo:.*?(\w+) : .*?PWD=([^\s]+) .*?COMMAND=(.*)", line
                )
                if sudo_match:
                    self.process_sudo_line(sudo_match, line, date_time_str, sudo_usage)
                elif not re.search(r"session (opened|closed) for user", line):
                    content = line[date_match.end():].strip()
                    other_activities.append((date_time_str, content))

                command_match = re.search(r"app-1 (\w+)", line)
                if command_match:
                    command = command_match.group(1)
                    if command not in logs_by_command:
                        logs_by_command[command] = []
                    logs_by_command[command].append((date_time_str, line.strip()))

        # Return all data structures including logs_by_command for new logic.
        return attacks, sudo_usage, other_activities, logs_by_command


    # Process a line related to sshd activity.
    def process_sshd_line(self, ip_match, line, date_time, attacks):
        ip = ip_match.group(1)
        port_match = re.search(r"port (\d+)", line)
        port = port_match.group(1) if port_match else "N/A"

        user = "N/A"
        user_match = re.search(r"for (\w+) from", line)
        invalid_user_match = re.search(r"invalid user (\w+)", line)

        if user_match:
            user = user_match.group(1)
        elif invalid_user_match:
            user = invalid_user_match.group(1)

        # Initialize dictionary structure for new IP addresses
        if ip not in attacks:
            attacks[ip] = {
                "start": [date_time],
                "end": [date_time],
                "success": 0,
                "fail": 0,
                "users": set(),
                "invalid_users": set(),
                "ports": set(),
                "success_details": [],
            }
        else:
            # Update start and end times for existing entries
            attacks[ip]["start"].append(date_time)
            attacks[ip]["end"].append(date_time)

        # Add port and update user details
        attacks[ip]["ports"].add(port)
        attacks[ip]["users"].add(user)
        if invalid_user_match:
            attacks[ip]["invalid_users"].add(user)

        # Record details based on line content
        if "Failed password" in line:
            attacks[ip]["fail"] += 1
        elif "Accepted password" in line:
            attacks[ip]["success"] += 1
            connection_detail = f"{date_time.strftime('%Y-%m-%d %H:%M:%S')}, Port: {port}, IP: {ip}, User: {user}"
            attacks[ip]["success_details"].append(connection_detail)

    # Process a line related to sudo usage.
    def process_sudo_line(self, sudo_match, line, date_time_str, sudo_usage):
        sudo_user = sudo_match.group(1)
        pwd = sudo_match.group(2)
        sudo_command = sudo_match.group(3).strip()
        sudo_usage.setdefault(sudo_user, []).append(
            {"date": date_time_str, "pwd": pwd, "command": sudo_command}
        )

    # Adds a styled table to the Excel sheet.
    def apply_table_style(self, sheet):
        # Adds and styles a table for better readability in Excel.
        table_name = f"{sheet.title.replace(' ', '_')}Table"
        table_ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        tab = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style
        sheet.add_table(tab)

    # Exports data to an Excel file.
    def export_to_excel(self, attacks, sudo_usage, other_activities, logs_by_command, file_name):
        # Creates workbook and sheets, then populates them with data.
        # Applies table styling and saves the workbook.
        wb = Workbook()
        ws = wb.active
        ws.title = "Attack Report"

        # Fill in the Attack Report sheet.
        ws.append(self.headers)
        for ip, data in attacks.items():
            domain_name = self.domain_name_cache.get(ip, "N/A")
            geo_info = self.geolocation_cache.get(ip, {
                "country": "N/A", "region": "N/A", "city": "N/A",
                "loc": "N/A", "org": "N/A", "postal": "N/A", "timezone": "N/A"
            })

            total_attempts = data["success"] + data["fail"]
            malicious_label = "No" if data["success"] > data["fail"] else "Not Sure"
            if total_attempts >= self.malicious_threshold and data["fail"] > 0:
                failure_rate = data["fail"] / total_attempts
                malicious_label = "Yes" if failure_rate > 0.9 else "Not Sure"

            row = [ip, domain_name] + list(geo_info.values()) + [
                min(data["start"]).strftime("%Y-%m-%d %H:%M:%S"),
                max(data["end"]).strftime("%Y-%m-%d %H:%M:%S"),
                data["success"], data["fail"], total_attempts, malicious_label,
                ", ".join(data["users"]), ", ".join(data["invalid_users"]),
                ", ".join(data["ports"]),
                "\n".join(data["success_details"]) if data["success_details"] else "N/A"
            ]
            ws.append(row)

        self.apply_table_style(ws)

        # Fill in the Sudo Usage sheet.
        sudo_ws = wb.create_sheet("Sudo Usage")
        sudo_ws.append(["User", "Date", "PWD", "Command"])
        for user, commands in sudo_usage.items():
            for cmd in commands:
                sudo_ws.append([user, cmd["date"], cmd["pwd"], cmd["command"]])

        self.apply_table_style(sudo_ws)

        # Fill in the Other Activities sheet.
        other_ws = wb.create_sheet("Other Activities")
        other_ws.append(["Date", "Content"])
        for activity in other_activities:
            other_ws.append(activity)

        self.apply_table_style(other_ws)

        # Create a sheet for each command type identified after "app-1".
        for command, logs in logs_by_command.items():
            command_ws = wb.create_sheet(title=command.capitalize())
            command_ws.append(["Date", "Log"])
            for log_entry in logs:
                command_ws.append(log_entry)

            self.apply_table_style(command_ws)

        wb.save(filename=file_name)

    def export_to_csv(self, attacks, sudo_usage, other_activities, file_name):
        with open(file_name, mode="w", newline="") as file:
            writer = csv.writer(file)

            # Attack Report
            writer.writerow(self.headers)
            attack_rows = []  # Initialize a list to store attack data rows.
            for ip, data in attacks.items():
                domain_name = self.domain_name_cache.get(ip, "N/A")
                geo_info = self.geolocation_cache.get(
                    ip,
                    {
                        "country": "N/A",
                        "region": "N/A",
                        "city": "N/A",
                        "loc": "N/A",
                        "org": "N/A",
                        "postal": "N/A",
                        "timezone": "N/A",
                    },
                )

                total_attempts = data["success"] + data["fail"]
                malicious_label = (
                    "Yes"
                    if total_attempts >= self.malicious_threshold and data["fail"] > 0
                    else "No"
                )

                attack_rows.append(
                    [
                        ip,
                        domain_name,
                        geo_info["country"],
                        geo_info["region"],
                        geo_info["city"],
                        geo_info["loc"],
                        geo_info["org"],
                        geo_info["postal"],
                        geo_info["timezone"],
                        min(data["start"]).strftime("%Y-%m-%d %H:%M:%S"),
                        max(data["end"]).strftime("%Y-%m-%d %H:%M:%S"),
                        data["success"],
                        data["fail"],
                        total_attempts,
                        malicious_label,
                        ", ".join(data["users"]),
                        ", ".join(data["invalid_users"]),
                        ", ".join(data["ports"]),
                        (
                            "; ".join(data.get("success_details", []))
                            if data["success"] < 6 or malicious_label == "Yes"
                            else "N/A"
                        ),
                    ]
                )
            writer.writerows(attack_rows)

            # Sudo Usage
            writer.writerow([])
            writer.writerow(["User", "Date", "PWD", "Command"])
            sudo_rows = [
                [user, cmd_info["date"], cmd_info["pwd"], cmd_info["command"]]
                for user, commands in sudo_usage.items()
                for cmd_info in commands
            ]
            writer.writerows(sudo_rows)

            # Other Activities
            writer.writerow([])
            writer.writerow(["Date", "Content"])
            other_activity_rows = [
                [date_str, content] for date_str, content in other_activities
            ]
            writer.writerows(other_activity_rows)


import asyncio
import sys

if __name__ == "__main__":
    # Default export format is Excel.
    export_format = "xlsx"

    # Check if at least one argument (the log file) is provided.
    if len(sys.argv) < 2:
        print("Usage: python3 test2.py <log_file> [-csv]")
        sys.exit(1)

    # The first argument after the script name is assumed to be the log file.
    log_file_arg = sys.argv[1]

    # Check if there is a request to export as CSV.
    if len(sys.argv) > 2 and sys.argv[2] == "-csv":
        export_format = "csv"

    # Initialize the parser with the log file.
    parser = AuthLogParser(log_file_arg)

    # Parse the log file.
    attacks_data, sudo_usage, other_activities, logs_by_command = parser.parse_auth_log()

    # Resolve IP addresses for the attacks data.
    ip_addresses = list(attacks_data.keys())
    asyncio.run(parser.resolve_addresses_batched(ip_addresses))

    # Export the parsed data based on the specified format.
    if export_format == "csv":
        print("CSV export functionality is currently limited. Exporting as Excel instead.")
        export_format = "xlsx"

    if export_format == "xlsx":
        output_file_name = "attacks_report.xlsx"
        parser.export_to_excel(
            attacks_data, sudo_usage, other_activities, logs_by_command, output_file_name
        )
        print(f"Report saved to {output_file_name}.")
