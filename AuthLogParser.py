import os  # Pour les fonctionnalités du système d'exploitation
import re  # Pour les expressions régulières
import csv  # Pour la manipulation des fichiers CSV
import sys  # Pour les fonctionnalités système
import asyncio  # Pour la programmation asynchrone
import aiohttp  # Pour les requêtes HTTP asynchrones
import socket  # Pour les opérations réseau
from datetime import datetime  # Pour manipuler les dates et heures
from openpyxl import Workbook  # Pour créer des fichiers Excel
from openpyxl.worksheet.table import (  # Pour créer des tableaux dans Excel
    Table,
    TableStyleInfo,
)
from openpyxl.utils import (  # Pour obtenir la lettre de la colonne Excel par numéro
    get_column_letter,
)
from tqdm import tqdm  # Pour afficher des barres de progression


# Classe pour analyser les journaux d'authentification
class AuthLogParser:
    def __init__(self, log_file):
        # Initialisation avec le fichier de journal
        self.log_file = log_file
        # Cache pour les noms de domaine associés aux adresses IP
        self.domain_name_cache = {}
        # Cache pour les informations de géolocalisation associées aux adresses IP
        self.geolocation_cache = {}
        # Réseaux locaux connus
        self.local_networks = [
            ("10.0.0.0", "10.255.255.255"),
            ("172.16.0.0", "172.31.255.255"),
            ("192.168.0.0", "192.168.255.255"),
        ]
        # En-têtes pour les données d'attaque
        self.headers = [
            "IP Address",
            "Domain Name",
            "Country",
            "Region",
            "City",
            "Localisation",
            "Organization",
            "Postal",
            "Timezone",
            "Start Time",
            "End Time",
            "Successful Attempts",
            "Failed Attempts",
            "Total Attempts",
            "Success/failed",
            "Malicious/Not Sure/No",
            "Impacted Users",
            "Invalid Users",
            "Ports",
            "Success Details",
        ]
        # Seuil pour déterminer si une activité est malveillante
        self.malicious_threshold = 5
        # Taille du lot pour la résolution asynchrone des adresses IP
        self.batch_size = 100
        # Temps d'attente pour les requêtes réseau
        self.timeout = 5
        # Étiquette pour les activités non sûres
        self.Not_sure = "Not Sure"

    # Vérifie si une adresse IP est locale
    def is_local_ip(self, ip_address):
        ip_int = self.ip_to_int(ip_address)
        return any(
            ip_int >= self.ip_to_int(start) and ip_int <= self.ip_to_int(end)
            for start, end in self.local_networks
        )

    # Convertit une adresse IP en entier
    def ip_to_int(self, ip):
        parts = ip.split(".")
        return (
            (int(parts[0]) << 24)
            | (int(parts[1]) << 16)
            | (int(parts[2]) << 8)
            | int(parts[3])
        )

    # Résolution asynchrone du nom de domaine associé à une adresse IP
    async def get_domain_name(self, ip_address):
        # Vérifie si l'adresse IP est déjà en cache ou si elle est locale
        if ip_address in self.domain_name_cache or self.is_local_ip(ip_address):
            return
        try:
            # Résolution du nom de domaine avec un délai d'attente
            hostname, _, _ = await asyncio.wait_for(
                asyncio.get_event_loop().run_in_executor(
                    None, socket.gethostbyaddr, ip_address
                ),
                timeout=self.timeout,
            )
            self.domain_name_cache[ip_address] = hostname
        except Exception:
            # En cas d'erreur, marquer comme non disponible
            self.domain_name_cache[ip_address] = "N/A"

    # Résolution asynchrone de la géolocalisation associée à une adresse IP
    async def geolocate_ip(self, ip_address, session):
        # Vérifie si les informations de géolocalisation sont déjà en cache ou si l'adresse IP est locale
        if ip_address in self.geolocation_cache or self.is_local_ip(ip_address):
            return
        try:
            # Requête HTTP asynchrone pour obtenir les données de géolocalisation avec un délai d'attente
            async with session.get(
                f"https://ipinfo.io/{ip_address}/json", timeout=self.timeout
            ) as response:
                data = await response.json()
                # Enregistrement des informations de géolocalisation
                self.geolocation_cache[ip_address] = {
                    "country": data.get("country", "N/A"),
                    "region": data.get("region", "N/A"),
                    "city": data.get("city", "N/A"),
                    "loc": data.get("loc", "N/A"),
                    "org": data.get("org", "N/A"),
                    "postal": data.get("postal", "N/A"),
                    "timezone": data.get("timezone", "N/A"),
                }
        except Exception as e:
            # En cas d'erreur, marquer comme non disponible avec le message d'erreur
            self.geolocation_cache[ip_address] = {
                "country": "N/A",
                "region": "N/A",
                "city": "N/A",
                "loc": "N/A",
                "org": "N/A",
                "postal": "N/A",
                "timezone": "N/A",
                "error": str(e),
            }

    # Résolution asynchrone des adresses IP en lots
    async def resolve_addresses_batched(self, ip_addresses):
        async with aiohttp.ClientSession() as session:
            # Initialisation de la barre de progression en dehors de la boucle
            print("")
            with tqdm(
                total=len(ip_addresses), desc="Resolving IP Addresses", unit=" IP"
            ) as progress_bar:
                for i in range(0, len(ip_addresses), self.batch_size):
                    batch = ip_addresses[i : i + self.batch_size]
                    tasks = []
                    for ip_address in batch:
                        if (
                            ip_address not in self.domain_name_cache
                            and not self.is_local_ip(ip_address)
                        ):
                            tasks.append(self.get_domain_name(ip_address))
                        if (
                            ip_address not in self.geolocation_cache
                            and not self.is_local_ip(ip_address)
                        ):
                            tasks.append(self.geolocate_ip(ip_address, session))
                    # Exécution asynchrone des tâches pour le lot courant
                    await asyncio.gather(*tasks)
                    # Mise à jour de la barre de progression après chaque lot traité
                    progress_bar.update(len(batch))

    # Analyse du journal d'authentification pour les attaques et les commandes
    def parse_auth_log(self):
        # Dictionnaire pour stocker les données des attaques
        attacks = {}
        # Dictionnaire pour stocker les logs par commande
        logs_by_command = {}
        # Année actuelle pour les logs sans indication d'année
        current_year = datetime.now().year

        # Lecture du fichier de journal ligne par ligne
        with open(self.log_file, "r") as file:
            lines = file.readlines()  # Lisez toutes les lignes en une fois
            print("")
            # Itération sur chaque ligne du journal
            for line in tqdm(
                lines, desc="Analyzing log", unit=" lines"
            ):  # Ajoutez tqdm ici
                # Recherche de la date dans le format spécifique
                date_match = re.search(r"^\w{3} \d{1,2} \d{2}:\d{2}:\d{2}", line)
                if not date_match:
                    continue

                # Construction de la date avec l'année actuelle
                date_str = f"{date_match.group(0)} {current_year}"
                date_time = datetime.strptime(date_str, "%b %d %H:%M:%S %Y")
                date_time_str = date_time.strftime("%Y-%m-%d %H:%M:%S")

                # Recherche du PID dans la ligne
                pid_match = re.search(r"\[(\d+)\]", line)
                pid = pid_match.group(1) if pid_match else "N/A"

                # Recherche de l'adresse IP dans la ligne
                ip_match = re.search(r"from (\d+\.\d+\.\d+\.\d+)", line)
                if ip_match:
                    self.process_sshd_line(ip_match, line, date_time, attacks)

                # Recherche de la commande dans la ligne
                command_match = re.search(
                    r"\w{3} \d{1,2} \d{2}:\d{2}:\d{2} \S+ (\w+)", line
                )

                if command_match:
                    command = command_match.group(1)
                    if command not in logs_by_command:
                        logs_by_command[command] = []
                    logs_by_command[command].append((date_time_str, pid, line.strip()))

        return attacks, logs_by_command

    # Traitement d'une ligne du journal SSH
    def process_sshd_line(self, ip_match, line, date_time, attacks):
        # Extraction de l'adresse IP à partir de la correspondance
        ip = ip_match.group(1)
        # Recherche du numéro de port dans la ligne
        port_match = re.search(r"port (\d+)", line)
        port = port_match.group(1) if port_match else "N/A"

        # Initialisation de l'utilisateur comme inconnu
        user = "N/A"
        # Recherche du nom d'utilisateur dans la ligne pour les authentifications réussies
        user_match = re.search(r"for (\w+) from", line)
        # Recherche du nom d'utilisateur dans la ligne pour les authentifications échouées
        invalid_user_match = re.search(r"invalid user (\w+)", line)

        # Si un nom d'utilisateur est trouvé, l'assigner à la variable 'user'
        if user_match:
            user = user_match.group(1)
        elif invalid_user_match:
            user = invalid_user_match.group(1)

        # Si l'adresse IP n'est pas déjà enregistrée comme attaque, initialiser ses données
        if ip not in attacks:
            attacks[ip] = {
                "start": [date_time],
                "end": [date_time],
                "success": 0,
                "fail": 0,
                "key_auth": 0,  # Ajout pour les authentifications par clé
                "users": set(),
                "invalid_users": set(),
                "ports": set(),
                "success_details": [],
            }
        else:
            # Si l'adresse IP est déjà enregistrée, mettre à jour les informations
            attacks[ip]["start"].append(date_time)
            attacks[ip]["end"].append(date_time)

        # Ajout du port utilisé à l'ensemble des ports pour cette adresse IP
        attacks[ip]["ports"].add(port)
        # Ajout du nom d'utilisateur à l'ensemble des utilisateurs pour cette adresse IP
        attacks[ip]["users"].add(user)
        # Si un nom d'utilisateur invalide est trouvé, l'ajouter à l'ensemble des utilisateurs invalides
        if invalid_user_match:
            attacks[ip]["invalid_users"].add(user)

        # Traitement des différentes lignes en fonction du type de connexion
        if "Failed password" in line:
            # Incrémenter le nombre d'authentifications échouées
            attacks[ip]["fail"] += 1
        elif "Accepted password" in line or "Accepted publickey" in line:
            # Incrémenter le nombre d'authentifications réussies
            attacks[ip]["success"] += 1
            # Si l'authentification a utilisé une clé publique, incrémenter le compteur d'authentifications par clé
            if "Accepted publickey" in line:
                attacks[ip]["key_auth"] += 1
            # Détails de la connexion réussie
            connection_detail = (
                f"{date_time.strftime('%Y-%m-%d %H:%M:%S')}, Port: {port}, User: {user}"
            )
            attacks[ip]["success_details"].append(connection_detail)

    def apply_table_style(self, sheet):
        # Création du nom et de la référence de la table
        table_name = f"{sheet.title.replace(' ', '_')}Table"
        table_ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        # Création de l'objet Table avec le nom et la référence spécifiés
        tab = Table(displayName=table_name, ref=table_ref)
        # Définition du style de la table
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style
        # Ajout de la table à la feuille
        sheet.add_table(tab)

    # Exporter les données vers un fichier Excel
    def export_to_excel(self, attacks, logs_by_command, file_name):
        # Création d'un classeur Excel
        wb = Workbook()
        # Sélection de la feuille active
        ws = wb.active
        # Nom de la feuille
        ws.title = "Attack Report"

        # Ajout des en-têtes à la feuille Excel
        ws.append(self.headers)
        print("")
        # Itération sur chaque adresse IP dans les données d'attaque
        for ip, data in tqdm(
            attacks.items(), desc="Exporting attacks to Excel", unit="row"
        ):
            # Récupération du nom de domaine associé à l'adresse IP
            domain_name = self.domain_name_cache.get(ip, "N/A")
            # Récupération des informations de géolocalisation associées à l'adresse IP
            geo_info = self.geolocation_cache.get(
                ip,
                {
                    "country": "N/A",
                    "region": "N/A",
                    "city": "N/A",
                    "loc": "N/A",
                    "org": "N/A",
                    "postal": "N/A",
                    "timezone": "N/A",
                },
            )

            # Calcul du nombre total de tentatives d'authentification
            total_attempts = data["success"] + data["fail"]
            # Calcul du ratio de réussite/échec des tentatives d'authentification
            success_fail_ratio = (
                str(data["success"] / data["fail"])
                if data["fail"]
                else "N/A" if data["success"] == 0 else "1"
            )

            # Détermination de l'étiquette de malveillance en fonction des données d'attaque
            if data["success"] > data["fail"]:
                malicious_label = "No"
            elif data["fail"] == 0 or total_attempts < self.malicious_threshold:
                malicious_label = self.Not_sure
            else:
                malicious_label = (
                    "Yes" if data["fail"] / total_attempts > 0.9 else self.Not_sure
                )

            # Construction de la ligne de données pour l'adresse IP actuelle
            row = [
                ip,
                domain_name,
                geo_info.get("country", "N/A"),
                geo_info.get("region", "N/A"),
                geo_info.get("city", "N/A"),
                geo_info.get("loc", "N/A"),
                geo_info.get("org", "N/A"),
                geo_info.get("postal", "N/A"),
                geo_info.get("timezone", "N/A"),
                min(data["start"]).strftime("%Y-%m-%d %H:%M:%S"),
                max(data["end"]).strftime("%Y-%m-%d %H:%M:%S"),
                data["success"],
                data["fail"],
                total_attempts,
                success_fail_ratio,
                malicious_label,
                ", ".join(data["users"]),
                ", ".join(data["invalid_users"]),
                ", ".join(data["ports"]),
                "; ".join(data["success_details"]),
            ]
            # Ajout de la ligne de données à la feuille Excel
            ws.append(row)

        # Application du style de tableau à la feuille Excel
        self.apply_table_style(ws)

        # Création d'une feuille pour les commandes uniques
        unique_commands_ws = wb.create_sheet(title="Unique Commands")

        # Ajout des en-têtes pour la feuille des commandes uniques
        unique_commands_ws.append(["Command", "Occurrences", "Go to Attack Report"])

        # Comptage des occurrences de chaque commande
        command_counts = {}
        for command, logs in logs_by_command.items():
            command_counts[command] = len(logs)

        # Ajout des données des commandes et de leurs occurrences dans la feuille
        print("")
        # Itération sur les données des commandes
        for command, count in tqdm(
            logs_by_command.items(), desc="Processing commands", unit="cmd"
        ):
            # Ajout d'une ligne pour chaque commande et son nombre d'occurrences
            unique_commands_ws.append([command, len(count), "Go to Attack Report"])

        # Ajout de l'hyperlien pour diriger vers la feuille "Attack Report" sur chaque ligne de commandes uniques
        print("")
        for row in unique_commands_ws.iter_rows(
            min_row=2, max_row=unique_commands_ws.max_row, min_col=3, max_col=3
        ):
            for cell in row:
                cell.hyperlink = f"#'Attack Report'!A1"

        # Application du style de tableau à la feuille des commandes uniques
        self.apply_table_style(unique_commands_ws)

        # Création d'une feuille pour chaque commande avec ses logs
        for command, logs in logs_by_command.items():
            # Création d'une nouvelle feuille pour chaque commande
            command_ws = wb.create_sheet(title=command.capitalize())
            # Ajout des en-têtes pour les logs de commande
            command_ws.append(["Date", "PID", "Log"])
            print("")
            # Itération sur les logs de la commande actuelle
            for log_entry in tqdm(logs, desc=f"Exporting {command} logs", unit="log"):
                # Ajout de chaque entrée de log comme une ligne dans la feuille
                command_ws.append(log_entry)

            # Application du style de tableau à la feuille de commande
            self.apply_table_style(command_ws)

        # Enregistrement du classeur Excel
        wb.save(filename=file_name)


if __name__ == "__main__":
    # Vérifie si un argument de fichier journal est fourni
    if len(sys.argv) < 2:
        print("Usage: python3 script.py <log_file>")
        sys.exit(1)

    # Récupère le chemin du fichier journal à partir des arguments
    log_file_arg = sys.argv[1]

    # Crée une instance de AuthLogParser avec le fichier journal spécifié
    parser = AuthLogParser(log_file_arg)

    # Analyse le fichier journal pour détecter les attaques et les commandes
    attacks_data, logs_by_command = parser.parse_auth_log()

    # Résout les adresses IP pour enrichir les données d'attaque
    asyncio.run(parser.resolve_addresses_batched(list(attacks_data.keys())))

    # Détermine le format du fichier de sortie (CSV/Excel) en fonction de l'option de l'utilisateur
    output_folder = "reports"
    os.makedirs(output_folder, exist_ok=True)  # Crée le dossier s'il n'existe pas

    if len(sys.argv) > 2 and sys.argv[2].lower() == "-csv":
        # Exporte les données vers un fichier CSV pour chaque commande
        for command, logs in logs_by_command.items():
            output_file_name = f"{output_folder}/attacks_report_{command}.csv"
            parser.export_to_csv(attacks_data, {command: logs}, output_file_name)
            print(f"Rapport CSV enregistré sous {output_file_name}.")
    else:
        # Exporte les données vers un fichier Excel
        output_file_name = f"{output_folder}/attacks_report.xlsx"
        parser.export_to_excel(attacks_data, logs_by_command, output_file_name)
        print(f"Rapport Excel enregistré sous {output_file_name}.")
