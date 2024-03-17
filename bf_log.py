# Script to parse authentication logs and export relevant data to an Excel file.

import re
import sys
import asyncio
import aiohttp
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Initialize caches for domain names and geolocation information.
domain_name_cache = {}
geolocation_cache = {}

# Check if the IP address is within local network ranges.
def is_local_ip(ip_address):
    local_networks = [
        ('10.0.0.0', '10.255.255.255'),
        ('172.16.0.0', '172.31.255.255'),
        ('192.168.0.0', '192.168.255.255'),
    ]
    ip_int = ip_to_int(ip_address)
    return any(ip_int >= ip_to_int(start) and ip_int <= ip_to_int(end) for start, end in local_networks)

# Convert IP address string to an integer.
def ip_to_int(ip):
    parts = ip.split('.')
    return (int(parts[0]) << 24) | (int(parts[1]) << 16) | (int(parts[2]) << 8) | int(parts[3])

# Asynchronously get the domain name for an IP address.
async def get_domain_name(ip_address):
    if ip_address in domain_name_cache or is_local_ip(ip_address):
        return
    try:
        info = await asyncio.get_event_loop().getaddrinfo(ip_address, None)
        domain_name_cache[ip_address] = info[0][4][0]
    except Exception:
        domain_name_cache[ip_address] = None

# Asynchronously geolocate an IP address using the ipinfo.io API.
async def geolocate_ip(ip_address, session):
    if ip_address in geolocation_cache or is_local_ip(ip_address):
        return
    try:
        async with session.get(f'https://ipinfo.io/{ip_address}/json', timeout=5) as response:
            data = await response.json()
            geolocation_cache[ip_address] = data
    except Exception as e:
        geolocation_cache[ip_address] = {'error': str(e)}

# Resolve domain names and geolocation information for a list of IP addresses.
async def resolve_addresses(ip_addresses):
    async with aiohttp.ClientSession() as session:
        tasks = [get_domain_name(ip) for ip in ip_addresses] + [geolocate_ip(ip, session) for ip in ip_addresses]
        await asyncio.gather(*tasks)

# Parse the authentication log file to extract attack, sudo usage, and other activity information.
def parse_auth_log(log_file):
    attacks = {}
    sudo_usage = {}
    other_activities = []
    current_year = datetime.now().year  # Use current year for log entries.

    with open(log_file, 'r') as file:
        for line in file:
            # Skip over cron session lines without significant content after the session action.
            cron_match = re.search(r'session (opened|closed) for user', line)
            if cron_match and not line[cron_match.end():].strip():
                continue

            # Extract the date and time from the beginning of each log line.
            date_match = re.search(r'^\w{3} \d{1,2} \d{2}:\d{2}:\d{2}', line)
            if not date_match:
                continue  # Skip lines without a date.

            # Append the current year to the log date and convert to datetime object.
            date_str = f"{date_match.group(0)} {current_year}"
            date_time = datetime.strptime(date_str, '%b %d %H:%M:%S %Y')
            date_time_str = date_time.strftime('%Y-%m-%d %H:%M:%S')

            # Process sshd and sudo related lines using provided functions.
            ip_match = re.search(r'from (\d+\.\d+\.\d+\.\d+)', line)
            if ip_match:
                process_sshd_line(ip_match, line, date_time, attacks)

            sudo_match = re.search(r'sudo:.*?(\w+) : .*?PWD=([^\s]+).*?COMMAND=(.*)', line)
            if sudo_match:
                process_sudo_line(sudo_match, line, date_time_str, sudo_usage)

            # Collect other non-cron, non-sshd, non-sudo lines.
            elif not cron_match:
                content = line[date_match.end():].strip()  # Content after the date.
                other_activities.append((date_time_str, content))

    return attacks, sudo_usage, other_activities

# Process a line related to sshd activity.
def process_sshd_line(ip_match, line, date_time, attacks):
    ip = ip_match.group(1)
    port_match = re.search(r'port (\d+)', line)
    user_match = re.search(r'for (\w+) from', line)
    invalid_user_match = re.search(r'invalid user (\w+)', line)
    port = port_match.group(1) if port_match else 'N/A'

    attacks.setdefault(ip, {
        'start': [], 'end': [], 'success': 0, 'fail': 0, 'users': set(),
        'invalid_users': set(), 'ports': set()
    })
    attacks[ip]['start'].append(date_time)
    attacks[ip]['end'].append(date_time)
    attacks[ip]['ports'].add(port)

    if user_match:
        user = user_match.group(1)
        attacks[ip]['users'].add(user)
    if invalid_user_match:
        invalid_user = invalid_user_match.group(1)
        attacks[ip]['invalid_users'].add(invalid_user)

    if "Failed password" in line:
        attacks[ip]['fail'] += 1
    elif "Accepted password" in line:
        attacks[ip]['success'] += 1

# Process a line related to sudo usage.
def process_sudo_line(sudo_match, line, date_time_str, sudo_usage):
    sudo_user = sudo_match.group(1)
    pwd = sudo_match.group(2)
    sudo_command = sudo_match.group(3).strip()
    sudo_usage.setdefault(sudo_user, []).append({
        'date': date_time_str,
        'pwd': pwd,
        'command': sudo_command
    })

# Apply Openpyxl table styling for sorting features.
def apply_table_style(sheet):
    # Remplace les espaces par des underscores dans le nom de la table
    table_name = f"{sheet.title.replace(' ', '_')}Table"
    table_ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    tab = Table(displayName=table_name, ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    sheet.add_table(tab)

# Export collected data to an Excel file with sorting features.    
def export_to_excel(attacks, sudo_usage, other_activities, file_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attack Report"

    headers = ["IP Address", "Domain Name", "Country", "Region", "City", "Start Time", "End Time",
               "Successful Attempts", "Failed Attempts", "Total Attempts", "Success/Failure Ratio",
               "Impacted Users", "Invalid Users", "Ports", "Malicious"]
    ws.append(headers)

    for ip, data in attacks.items():
        domain_name = domain_name_cache.get(ip, 'N/A')
        geo_info = geolocation_cache.get(ip, {'country': 'N/A', 'region': 'N/A', 'city': 'N/A'})

        row = [
            ip, domain_name, geo_info.get('country', 'N/A'), geo_info.get('region', 'N/A'), geo_info.get('city', 'N/A'),
            min(data['start']).strftime('%Y-%m-%d %H:%M:%S'), max(data['end']).strftime('%Y-%m-%d %H:%M:%S'),
            data['success'], data['fail'], data['success'] + data['fail'],
            'Inf' if data['fail'] == 0 else round(data['success'] / data['fail'], 2),
            ', '.join(data['users']), ', '.join(data['invalid_users']), ', '.join(data['ports']),
            'Yes' if data['success'] < data['fail'] else 'No'
        ]
        ws.append(row)

    # Apply table formatting to the attack report
    table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    tab = Table(displayName="AttackReportTable", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Sudo Usage Sheet
    sudo_ws = wb.create_sheet("Sudo Usage")
    sudo_ws.append(["User", "Date", "PWD", "Command"])
    for user, commands in sudo_usage.items():
        for command_info in commands:
            sudo_ws.append([
                user,
                command_info['date'],
                command_info['pwd'],
                command_info['command']
            ])

    # Appliquer le formatage de table avec tri sur "Sudo Usage"
    apply_table_style(sudo_ws)
    # Other Activities Sheet
    other_ws = wb.create_sheet("Other Activities")
    other_ws.append(["Date", "Content"])
    for date_str, content in other_activities:
        other_ws.append([date_str, content])
    # Appliquer le formatage de table avec tri sur "Other Activities"
    apply_table_style(other_ws)

    wb.save(filename=file_name)

if __name__ == "__main__":
     # Check for command-line arguments and parse log file.
    if len(sys.argv) < 2:
        print("Usage: python3 script.py log_file")
        sys.exit(1)

    log_file = sys.argv[1]
    attacks_data, sudo_usage, other_activities = parse_auth_log(log_file)

    # Resolve IP addresses before exporting to Excel
    ip_addresses = list(attacks_data.keys())
    asyncio.run(resolve_addresses(ip_addresses))

    export_to_excel(attacks_data, sudo_usage, other_activities, 'attacks_report.xlsx')
    print('Report saved to attacks_report.xlsx.')
